"""
SyncDB survey ingestion service.

Behavior:
- CSV uploads are ingested directly.
- If an Excel sheet is explicitly provided, only that sheet is processed.
- Otherwise, all workbook sheets are inspected, classified, and processed.
- Data sheets are combined into one dataset.
- Datamap and metadata sheets are preserved in dataset metadata.
"""

import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Optional

import pandas as pd
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.id_generator import generate_id
from app.utils.syncdb_helpers import parse_json_strings

_DEMOGRAPHIC_KEYWORDS = {
    "age", "gender", "sex", "location", "country", "state", "city",
    "education", "income", "occupation", "marital", "ethnicity",
    "race", "region", "district", "language", "nationality",
}
_OPEN_TEXT_THRESHOLD = 30
_SHEET_SCAN_LIMIT = 25
_RESPONSE_INSERT_BATCH_SIZE = 1000
_AUTO_SHEET_TOKENS = {"", "string", "all", "auto", "*", "none", "null", "undefined"}

_DATAMAP_NAME_HINTS = {"datamap", "codebook", "schema", "legend", "mapping", "map"}
_METADATA_NAME_HINTS = {"metadata", "summary", "overview", "intro", "cover", "readme", "notes"}
_DATA_HEADER_HINTS = {"record", "uuid", "date", "status", "qtime", "vos", "markers", "vlist"}
_DATAMAP_VALUE_HINTS = {
    "participant identifier",
    "record number",
    "open text response",
    "open numeric response",
    "single coded response",
}
_METADATA_VALUE_HINTS = {
    "survey description",
    "target audience",
    "main industries",
    "survey type",
    "no. of respondents",
    "tags",
}


def _normalize_sheet_selector(sheet_name: int | str | None) -> int | str | None:
    if sheet_name is None:
        return None
    if isinstance(sheet_name, int):
        return sheet_name
    value = str(sheet_name).strip()
    if value.lower() in _AUTO_SHEET_TOKENS:
        return None
    if re.fullmatch(r"-?\d+", value):
        return int(value)
    return value


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_long_text_row(values: list[str]) -> bool:
    non_empty = [value for value in values if value]
    return bool(non_empty) and len(non_empty) <= 2 and any(len(value) >= 80 for value in non_empty)


def _score_header_candidate(row_values: list[str], following_rows: list[list[str]]) -> float:
    non_empty = [value for value in row_values if value]
    if len(non_empty) < 2:
        return float("-inf")

    unique_ratio = len({value.lower() for value in non_empty}) / len(non_empty)
    numeric_ratio = sum(value.replace(".", "", 1).isdigit() for value in non_empty) / len(non_empty)
    avg_len = sum(len(value) for value in non_empty) / len(non_empty)

    next_row_density = 0.0
    if following_rows:
        next_non_empty = [value for value in following_rows[0] if value]
        next_row_density = len(next_non_empty) / max(len(non_empty), 1)

    header_hits = sum(1 for value in non_empty if value.lower() in _DATA_HEADER_HINTS)
    bracket_hits = sum(1 for value in non_empty if value.startswith("[") or value.endswith("]"))

    score = float(len(non_empty) * 4)
    score += unique_ratio * 8
    score += max(0.0, 4 - numeric_ratio * 10)
    score += min(next_row_density, 2.0) * 4
    score += min(header_hits * 2, 10)
    score += min(bracket_hits, 6)

    if avg_len > 40:
        score -= 8
    if _looks_like_long_text_row(row_values):
        score -= 12
    return score


def _analyze_sheet(sheet: Any, sheet_index: int) -> dict[str, Any]:
    sampled_rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_stringify_cell(cell) for cell in row]
        if any(values):
            sampled_rows.append((row_number, values))
        if len(sampled_rows) >= _SHEET_SCAN_LIMIT:
            break

    max_width = max((sum(1 for value in values if value) for _, values in sampled_rows), default=0)
    best_header_row = 0
    best_header_score = float("-inf")
    long_text_rows = 0
    datamap_marker_rows = 0
    metadata_marker_rows = 0

    for idx, (row_number, values) in enumerate(sampled_rows):
        if _looks_like_long_text_row(values):
            long_text_rows += 1
        flattened = " | ".join(value.lower() for value in values if value)
        if any(hint in flattened for hint in _DATAMAP_VALUE_HINTS) or any(value.startswith("[") for value in values if value):
            datamap_marker_rows += 1
        if any(hint in flattened for hint in _METADATA_VALUE_HINTS):
            metadata_marker_rows += 1

        following_rows = [row for _, row in sampled_rows[idx + 1 : idx + 3]]
        score = _score_header_candidate(values, following_rows)
        if score > best_header_score:
            best_header_score = score
            best_header_row = row_number - 1

    row_extent = sheet.max_row or len(sampled_rows)
    estimated_row_count = max(row_extent - best_header_row - 1, max(len(sampled_rows) - 1, 0))

    return {
        "sheet_name": sheet.title,
        "sheet_index": sheet_index,
        "header_row": max(best_header_row, 0),
        "header_score": round(best_header_score, 2),
        "estimated_row_count": estimated_row_count,
        "sampled_non_empty_rows": len(sampled_rows),
        "sampled_max_width": max_width,
        "long_text_rows": long_text_rows,
        "datamap_marker_rows": datamap_marker_rows,
        "metadata_marker_rows": metadata_marker_rows,
    }


def _classify_sheet(analysis: dict[str, Any]) -> str:
    title_lower = analysis["sheet_name"].lower().strip()
    max_width = analysis["sampled_max_width"]

    if analysis["sampled_non_empty_rows"] == 0 or max_width == 0:
        return "empty"

    if any(hint in title_lower for hint in _DATAMAP_NAME_HINTS) or analysis["datamap_marker_rows"] >= 2:
        return "datamap"

    if max_width >= 4 and (re.fullmatch(r"[a-z]+\d+", title_lower) or title_lower in {"data", "responses", "response"}):
        return "data"

    if max_width >= 6 and analysis["header_score"] >= 8:
        return "data"

    if any(hint in title_lower for hint in _METADATA_NAME_HINTS):
        return "metadata"

    if analysis["metadata_marker_rows"] >= 1:
        return "metadata"

    if max_width <= 3 or analysis["long_text_rows"] >= 2:
        return "metadata"

    return "data" if max_width >= 5 else "metadata"


def _build_workbook_plan(
    file_bytes: bytes,
    explicit_sheet_name: int | str | None,
    explicit_header_row: Optional[int],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel survey ingestion") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        plans: list[dict[str, Any]] = []
        normalized_sheet = _normalize_sheet_selector(explicit_sheet_name)

        if normalized_sheet is not None:
            if isinstance(normalized_sheet, int):
                try:
                    sheet = workbook.worksheets[normalized_sheet]
                except IndexError as exc:
                    raise ValueError(f"Excel sheet index {normalized_sheet} is out of range") from exc
            else:
                if normalized_sheet not in workbook.sheetnames:
                    raise ValueError(
                        f"Excel sheet '{normalized_sheet}' not found. Available sheets: {workbook.sheetnames}"
                    )
                sheet = workbook[normalized_sheet]

            analysis = _analyze_sheet(sheet, workbook.sheetnames.index(sheet.title))
            analysis["classification"] = _classify_sheet(analysis)
            analysis["header_row"] = explicit_header_row if explicit_header_row is not None else analysis["header_row"]
            analysis["ingest_role"] = "data"
            analysis["selected"] = True
            plans.append(analysis)
            metadata = {
                "selection_mode": "explicit_sheet",
                "requested_sheet": normalized_sheet,
            }
            return plans, metadata

        for idx, sheet in enumerate(workbook.worksheets):
            analysis = _analyze_sheet(sheet, idx)
            analysis["classification"] = _classify_sheet(analysis)
            analysis["header_row"] = explicit_header_row if explicit_header_row is not None else analysis["header_row"]
            analysis["selected"] = analysis["classification"] != "empty"
            analysis["ingest_role"] = analysis["classification"] if analysis["selected"] else "skipped"
            plans.append(analysis)

        data_plans = [plan for plan in plans if plan["ingest_role"] == "data"]
        if not data_plans:
            non_empty_plans = [plan for plan in plans if plan["classification"] != "empty"]
            if non_empty_plans:
                fallback = max(
                    non_empty_plans,
                    key=lambda item: (item["header_score"], item["sampled_max_width"], item["estimated_row_count"]),
                )
                fallback["ingest_role"] = "data"
                fallback["classification"] = "data"

        metadata = {
            "selection_mode": "all_sheets",
            "requested_sheet": None,
        }
        return plans, metadata
    finally:
        workbook.close()


def _normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    columns: list[str] = []
    used: dict[str, int] = {}
    for index, column in enumerate(df.columns, start=1):
        name = _stringify_cell(column)
        if not name or name.lower().startswith("unnamed:"):
            name = f"column_{index}"
        suffix = used.get(name, 0)
        used[name] = suffix + 1
        if suffix:
            name = f"{name}_{suffix + 1}"
        columns.append(name)

    normalized = df.copy()
    normalized.columns = columns
    return normalized


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.dropna(how="all").dropna(axis=1, how="all")
    if cleaned.empty or len(cleaned.columns) == 0:
        return cleaned
    return _normalize_dataframe_columns(cleaned)


def _read_excel_sheet(file_bytes: bytes, sheet_name: int | str, header_row: Optional[int]) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row)


def _read_excel_sheet_raw(file_bytes: bytes, sheet_name: int | str) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)


def _serialize_non_data_sheet(df: pd.DataFrame) -> list[list[Any]]:
    cleaned = df.dropna(how="all").dropna(axis=1, how="all")
    if cleaned.empty or len(cleaned.columns) == 0:
        return []
    return json.loads(cleaned.to_json(orient="values"))


def _infer_column_type_from_stats(col_name: str, stats: dict[str, Any]) -> str:
    col_lower = col_name.lower().strip()
    if any(keyword in col_lower for keyword in _DEMOGRAPHIC_KEYWORDS):
        return "demographic"
    if stats["is_open_text"]:
        return "open_text"
    return "single_choice"


def _update_schema_stats(
    df: pd.DataFrame,
    column_order: list[str],
    column_stats: dict[str, dict[str, Any]],
) -> None:
    for column in df.columns:
        if column not in column_stats:
            column_order.append(column)
            column_stats[column] = {
                "is_open_text": False,
                "counts": Counter(),
                "unique_values": set(),
            }

        stats = column_stats[column]
        values = df[column].dropna().astype(str)
        if values.empty:
            continue

        if stats["is_open_text"]:
            continue

        value_counts = values.value_counts()
        col_lower = column.lower().strip()
        is_demographic = any(keyword in col_lower for keyword in _DEMOGRAPHIC_KEYWORDS)

        if not is_demographic:
            stats["unique_values"].update(value_counts.index.tolist())
            if len(stats["unique_values"]) > _OPEN_TEXT_THRESHOLD:
                stats["is_open_text"] = True
                stats["counts"] = Counter()
                stats["unique_values"] = set()
                continue

        for value, count in value_counts.items():
            stats["counts"][value] += int(count)


def _build_question_schema(column_order: list[str], column_stats: dict[str, dict[str, Any]]) -> list[dict]:
    schema = []
    for column in column_order:
        stats = column_stats[column]
        col_type = _infer_column_type_from_stats(column, stats)
        options = sorted(stats["counts"].keys()) if col_type in {"single_choice", "demographic"} else []
        schema.append({"column": column, "text": column, "options": options, "type": col_type})
    return schema


def _build_aggregation(question_schema: list[dict], column_stats: dict[str, dict[str, Any]]) -> dict:
    results: dict[str, list[dict[str, Any]]] = {}
    for question in question_schema:
        if question["type"] == "open_text":
            continue
        counts = column_stats.get(question["column"], {}).get("counts", Counter())
        results[question["text"]] = [
            {"option": option, "count": int(count)}
            for option, count in counts.items()
        ]
    return results


async def _insert_response_batch(db: AsyncSession, batch: list[dict]) -> None:
    if not batch:
        return
    await db.execute(
        text(
            """
            INSERT INTO sync_survey.response
                (id, dataset_id, respondent_id, answers,
                 status, workspace_id, region, year, source_format, subject_key)
            VALUES
                (:id, :dataset_id, :respondent_id, CAST(:answers AS jsonb),
                 :status, :workspace_id, :region, :year, :source_format, :subject_key)
            """
        ),
        batch,
    )


async def _insert_data_sheet_responses(
    db: AsyncSession,
    file_bytes: bytes,
    dataset_id: str,
    data_plans: list[dict[str, Any]],
    question_schema: list[dict],
    workspace_id: Optional[str],
    region: Optional[str],
    year: Optional[int],
    filename: str,
    source_format: str,
    ingested_at: str,
) -> None:
    batch: list[dict] = []
    respondent_index = 0

    for plan in data_plans:
        df = _clean_dataframe(_read_excel_sheet(file_bytes, plan["sheet_name"], plan["header_row"]))
        if df.empty or len(df.columns) == 0:
            continue

        rows_as_dicts: list[dict] = json.loads(df.to_json(orient="records"))
        for row in rows_as_dicts:
            row = parse_json_strings(row)
            answers: dict[str, Any] = {}
            demographics: dict[str, str] = {}
            for question in question_schema:
                value = row.get(question["column"])
                if value is None:
                    continue
                if question["type"] == "demographic":
                    demographics[question["column"]] = str(value)
                else:
                    answers[question["column"]] = str(value)
            if demographics:
                answers["demographics"] = demographics

            respondent_index += 1
            envelope = {
                "source_type": "survey",
                "workspace_id": workspace_id,
                "region": region,
                "year": year,
                "source_file": filename,
                "source_format": source_format,
                "source_sheet": plan["sheet_name"],
                "ingested_at": ingested_at,
                "status": "pending",
                "payload": answers,
            }
            batch.append(
                {
                    "id": generate_id(),
                    "dataset_id": dataset_id,
                    "respondent_id": str(respondent_index),
                    "answers": json.dumps(envelope),
                    "status": "pending",
                    "workspace_id": workspace_id,
                    "region": region,
                    "year": year,
                    "source_format": source_format,
                    "subject_key": None,
                }
            )

            if len(batch) >= _RESPONSE_INSERT_BATCH_SIZE:
                await _insert_response_batch(db, batch)
                batch = []

    await _insert_response_batch(db, batch)


async def ingest_survey_csv(
    db: AsyncSession,
    file_bytes: bytes,
    filename: str,
    exploration_id: Optional[str],
    user_id: str,
    workspace_id: Optional[str],
    region: Optional[str],
    year: Optional[int],
    sheet_name: int | str | None = None,
    header_row: Optional[int] = None,
) -> dict:
    """
    Parse a survey file, classify workbook sheets, and ingest survey responses.
    """
    ext = ("." + filename.rsplit(".", 1)[-1]).lower() if "." in filename else ""
    source_format = "excel" if ext in (".xlsx", ".xls") else "csv"
    dataset_id = generate_id()
    ingested_at = datetime.now(timezone.utc).isoformat()

    if source_format == "csv":
        df = _clean_dataframe(pd.read_csv(io.BytesIO(file_bytes), header=header_row if header_row is not None else 0))
        if df.empty or len(df.columns) == 0:
            raise ValueError("No tabular survey data found after removing empty rows and columns")

        rows_as_dicts: list[dict] = json.loads(df.to_json(orient="records"))
        question_schema = []
        for column in df.columns:
            values = df[column].dropna().astype(str)
            col_type = "demographic" if any(k in column.lower() for k in _DEMOGRAPHIC_KEYWORDS) else (
                "open_text" if values.nunique() > _OPEN_TEXT_THRESHOLD else "single_choice"
            )
            options = sorted(values.unique().tolist()) if col_type in {"single_choice", "demographic"} else []
            question_schema.append({"column": column, "text": column, "options": options, "type": col_type})

        dataset_metadata = {
            "selection_mode": "csv",
            "requested_sheet": None,
            "processed_sheets": [],
            "data_sheets": [],
            "datamap_sheets": [],
            "metadata_sheets": [],
            "header_row": header_row if header_row is not None else 0,
            "column_count": int(len(df.columns)),
            "respondent_count": int(len(rows_as_dicts)),
        }

        await db.execute(
            text(
                """
                INSERT INTO sync_survey.dataset
                    (id, source_file, respondent_count,
                     question_schema, exploration_id, uploaded_by, metadata)
                VALUES
                    (:id, :source_file, :respondent_count,
                     CAST(:question_schema AS jsonb), :exploration_id, :uploaded_by,
                     CAST(:metadata AS jsonb))
                """
            ),
            {
                "id": dataset_id,
                "source_file": filename,
                "respondent_count": len(rows_as_dicts),
                "question_schema": json.dumps(question_schema),
                "exploration_id": exploration_id,
                "uploaded_by": user_id,
                "metadata": json.dumps(dataset_metadata),
            },
        )

        batch: list[dict] = []
        for index, row in enumerate(rows_as_dicts, start=1):
            row = parse_json_strings(row)
            answers: dict[str, Any] = {}
            demographics: dict[str, str] = {}
            for question in question_schema:
                value = row.get(question["column"])
                if value is None:
                    continue
                if question["type"] == "demographic":
                    demographics[question["column"]] = str(value)
                else:
                    answers[question["column"]] = str(value)
            if demographics:
                answers["demographics"] = demographics

            envelope = {
                "source_type": "survey",
                "workspace_id": workspace_id,
                "region": region,
                "year": year,
                "source_file": filename,
                "source_format": source_format,
                "ingested_at": ingested_at,
                "status": "pending",
                "payload": answers,
            }
            batch.append(
                {
                    "id": generate_id(),
                    "dataset_id": dataset_id,
                    "respondent_id": str(index),
                    "answers": json.dumps(envelope),
                    "status": "pending",
                    "workspace_id": workspace_id,
                    "region": region,
                    "year": year,
                    "source_format": source_format,
                    "subject_key": None,
                }
            )
            if len(batch) >= _RESPONSE_INSERT_BATCH_SIZE:
                await _insert_response_batch(db, batch)
                batch = []

        await _insert_response_batch(db, batch)

        agg_results = {
            question["text"]: [
                {"option": option, "count": int(count)}
                for option, count in df[question["column"]].dropna().astype(str).value_counts().items()
            ]
            for question in question_schema
            if question["type"] != "open_text"
        }

        await db.execute(
            text(
                """
                INSERT INTO sync_survey.aggregation (id, dataset_id, results)
                VALUES (:id, :dataset_id, CAST(:results AS jsonb))
                """
            ),
            {
                "id": generate_id(),
                "dataset_id": dataset_id,
                "results": json.dumps(agg_results),
            },
        )

        await db.commit()
        return await get_survey_dataset(db, dataset_id)

    workbook_plans, selection_metadata = _build_workbook_plan(
        file_bytes=file_bytes,
        explicit_sheet_name=sheet_name,
        explicit_header_row=header_row,
    )

    column_order: list[str] = []
    column_stats: dict[str, dict[str, Any]] = {}
    respondent_count = 0

    dataset_metadata = {
        **selection_metadata,
        "processed_sheets": [],
        "data_sheets": [],
        "datamap_sheets": [],
        "metadata_sheets": [],
    }

    data_plans: list[dict[str, Any]] = []

    for plan in workbook_plans:
        plan_summary = {
            "sheet_name": plan["sheet_name"],
            "sheet_index": plan["sheet_index"],
            "classification": plan["classification"],
            "ingest_role": plan["ingest_role"],
            "header_row": plan["header_row"],
            "header_score": plan["header_score"],
            "sampled_max_width": plan["sampled_max_width"],
            "estimated_row_count": plan["estimated_row_count"],
        }

        if plan["ingest_role"] == "data":
            df = _clean_dataframe(_read_excel_sheet(file_bytes, plan["sheet_name"], plan["header_row"]))
            if df.empty or len(df.columns) == 0:
                plan_summary["status"] = "skipped_empty_after_parse"
                dataset_metadata["processed_sheets"].append(plan_summary)
                continue

            row_count = int(len(df))
            column_count = int(len(df.columns))
            plan_summary["status"] = "processed"
            plan_summary["row_count"] = row_count
            plan_summary["column_count"] = column_count
            dataset_metadata["processed_sheets"].append(plan_summary)
            dataset_metadata["data_sheets"].append(plan_summary.copy())

            respondent_count += row_count
            _update_schema_stats(df, column_order, column_stats)
            data_plans.append(plan)
            continue

        raw_df = _read_excel_sheet_raw(file_bytes, plan["sheet_name"])
        raw_rows = _serialize_non_data_sheet(raw_df)
        if not raw_rows:
            plan_summary["status"] = "skipped_empty_after_parse"
            dataset_metadata["processed_sheets"].append(plan_summary)
            continue

        plan_summary["status"] = "processed"
        plan_summary["row_count"] = len(raw_rows)
        plan_summary["column_count"] = max((len(row) for row in raw_rows), default=0)
        dataset_metadata["processed_sheets"].append(plan_summary)

        payload = {
            "sheet_name": plan["sheet_name"],
            "sheet_index": plan["sheet_index"],
            "classification": plan["classification"],
            "header_row": plan["header_row"],
            "rows": raw_rows,
        }
        if plan["ingest_role"] == "datamap":
            dataset_metadata["datamap_sheets"].append(payload)
        else:
            dataset_metadata["metadata_sheets"].append(payload)

    if not data_plans or respondent_count == 0:
        raise ValueError("No survey response sheets were found in the workbook")

    question_schema = _build_question_schema(column_order, column_stats)
    agg_results = _build_aggregation(question_schema, column_stats)
    dataset_metadata["column_count"] = len(column_order)
    dataset_metadata["respondent_count"] = respondent_count

    await db.execute(
        text(
            """
            INSERT INTO sync_survey.dataset
                (id, source_file, respondent_count,
                 question_schema, exploration_id, uploaded_by, metadata)
            VALUES
                (:id, :source_file, :respondent_count,
                 CAST(:question_schema AS jsonb), :exploration_id, :uploaded_by,
                 CAST(:metadata AS jsonb))
            """
        ),
        {
            "id": dataset_id,
            "source_file": filename,
            "respondent_count": respondent_count,
            "question_schema": json.dumps(question_schema),
            "exploration_id": exploration_id,
            "uploaded_by": user_id,
            "metadata": json.dumps(dataset_metadata),
        },
    )

    await _insert_data_sheet_responses(
        db=db,
        file_bytes=file_bytes,
        dataset_id=dataset_id,
        data_plans=data_plans,
        question_schema=question_schema,
        workspace_id=workspace_id,
        region=region,
        year=year,
        filename=filename,
        source_format=source_format,
        ingested_at=ingested_at,
    )

    await db.execute(
        text(
            """
            INSERT INTO sync_survey.aggregation (id, dataset_id, results)
            VALUES (:id, :dataset_id, CAST(:results AS jsonb))
            """
        ),
        {
            "id": generate_id(),
            "dataset_id": dataset_id,
            "results": json.dumps(agg_results),
        },
    )

    await db.commit()
    return await get_survey_dataset(db, dataset_id)


async def get_survey_dataset(db: AsyncSession, dataset_id: str) -> Optional[dict]:
    row = await db.execute(
        text("SELECT * FROM sync_survey.dataset WHERE id = :id"),
        {"id": dataset_id},
    )
    result = row.mappings().first()
    return dict(result) if result else None


async def list_survey_datasets(
    db: AsyncSession,
    exploration_id: Optional[str] = None,
) -> list[dict]:
    if exploration_id:
        rows = await db.execute(
            text("SELECT * FROM sync_survey.dataset WHERE exploration_id = :eid ORDER BY uploaded_at DESC"),
            {"eid": exploration_id},
        )
    else:
        rows = await db.execute(text("SELECT * FROM sync_survey.dataset ORDER BY uploaded_at DESC"))
    return [dict(row) for row in rows.mappings().all()]


async def get_survey_aggregation(db: AsyncSession, dataset_id: str) -> Optional[dict]:
    row = await db.execute(
        text("SELECT * FROM sync_survey.aggregation WHERE dataset_id = :did"),
        {"did": dataset_id},
    )
    result = row.mappings().first()
    return dict(result) if result else None


async def link_exploration(db: AsyncSession, dataset_id: str, exploration_id: str) -> dict:
    await db.execute(
        text("UPDATE sync_survey.dataset SET exploration_id = :eid WHERE id = :did"),
        {"eid": exploration_id, "did": dataset_id},
    )
    await db.commit()
    return await get_survey_dataset(db, dataset_id)
